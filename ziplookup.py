import openpyxl
import njData
# import pprint

wb = openpyxl.load_workbook('nj.xlsx')
ws = wb['NJ']

#lb = openpyxl.load_workbook('lookup.xlsx')
#ls = lb['Lookup']
#lookupData = {}
#for row in range(2, ls.max_row + 1):
#    state = ls['A' + str(row)].value
#    zip = ls['B' + str(row)].value
#    terr = ls['E' + str(row)].value
#    lookupData.setdefault(state, {})
#    lookupData[state].setdefault(terr, {zip})

# resultFile = open('njData.py', 'w')
# resultFile.write('allData = ' + pprint.pformat(lookupData))
# resultFile.close

for rowNum in range(2, ws.max_row + 1):
    terr = ws['C' + str(rowNum)].value
    njZip = njData.allData[terr]
    print(njZip)
    ws['O' + str(rowNum)].value = njZip
    

wb.save('NJ.xlsx')