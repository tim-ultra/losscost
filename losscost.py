import xlrd 
import openpyxl 

# load input file
wb = xlrd.open_workbook("c:\losscost\iowa.xlsx")
ws = wb.sheet_by_name('Table 1')

# load output file
outfile = 'c:\losscost\output.xlsx'
nb = openpyxl.load_workbook(filename=outfile)
ns = nb['Output']

# function to get data from input file and write to output file
def codeSearch(x):
    for row in range(ws.nrows):
        for col in range(ws.ncols):
            if ws.cell_value(row, col) == x:
                a = ws.cell_value(row, col)
                b = ws.cell_value(row, col+1)
                c = ws.cell_value(row, col+2)
                nRow = ns.max_row + 1
                ns.cell(row=nRow, column=1, value=a)
                ns.cell(row=nRow, column=2, value=b)
                ns.cell(row=nRow, column=3, value=c)
                
# reads code list and searches for each one
data = xlrd.open_workbook("c:\losscost\codes.xlsx")
sheet = data.sheet_by_name('codes')
for row in range(sheet.nrows):
    codeSearch(sheet.cell_value(row,0))

# save output file    
nb.save(outfile)